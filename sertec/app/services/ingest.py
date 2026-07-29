"""Ingesta del Excel HISTORICO_SERTEC hacia la base de datos.

Lee las 3 hojas (Base SERTEC, Base SF, Resumen), valida el formato, crea la
`Carga` y guarda los snapshots de OST y SF. Al final dispara el motor de diffs.
"""
import datetime as dt
import re

import openpyxl
from sqlalchemy.orm import Session

from ..models import Carga, OstSnapshot, SfSnapshot
from . import sf_link

SHEET_SERTEC = "Base SERTEC"
SHEET_SF = "Base SF"
SHEET_RESUMEN = "Resumen"


class IngestError(Exception):
    pass


def _headers(ws):
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(h).strip(): i for i, h in enumerate(row) if h is not None}


def _as_dt(v):
    if isinstance(v, dt.datetime):
        return v
    if isinstance(v, dt.date):
        return dt.datetime(v.year, v.month, v.day)
    return None


def _as_int(v):
    try:
        if v is None or v == "":
            return None
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _as_float(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except (ValueError, TypeError):
        return None


def _s(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _parse_fecha_extraccion(ws_resumen) -> dt.datetime:
    """Lee 'Fecha extracción' de la hoja Resumen (formato dd-mm-YYYY HH:MM)."""
    for row in ws_resumen.iter_rows(values_only=True):
        if row and row[0] and "extrac" in str(row[0]).lower():
            raw = row[1]
            if isinstance(raw, dt.datetime):
                return raw
            m = re.search(r"(\d{2})-(\d{2})-(\d{4})[ T]+(\d{1,2}):(\d{2})", str(raw))
            if m:
                d, mo, y, h, mi = map(int, m.groups())
                return dt.datetime(y, mo, d, h, mi)
    raise IngestError("No se encontró 'Fecha extracción' en la hoja Resumen.")


def _parse_resumen_totales(ws_resumen) -> dict:
    tot = {}
    for row in ws_resumen.iter_rows(values_only=True):
        if row and row[0] is not None and row[1] is not None:
            tot[str(row[0]).strip()] = row[1]
    return tot


def parse_workbook_meta(path: str) -> dict:
    """Lee solo metadatos (fecha y totales) sin cargar los snapshots.

    Sirve para validar el archivo y detectar cargas duplicadas antes de insertar.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for req in (SHEET_SERTEC, SHEET_SF, SHEET_RESUMEN):
        if req not in wb.sheetnames:
            raise IngestError(f"Falta la hoja obligatoria '{req}' en el archivo.")
    fecha = _parse_fecha_extraccion(wb[SHEET_RESUMEN])
    totales = _parse_resumen_totales(wb[SHEET_RESUMEN])
    wb.close()
    return {"fecha_extraccion": fecha, "totales": totales}


def ingest_file(db: Session, path: str, archivo_nombre: str, usuario: str) -> Carga:
    """Procesa el Excel y crea la Carga con sus snapshots. Devuelve la Carga."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for req in (SHEET_SERTEC, SHEET_SF, SHEET_RESUMEN):
        if req not in wb.sheetnames:
            raise IngestError(f"Falta la hoja obligatoria '{req}' en el archivo.")

    fecha = _parse_fecha_extraccion(wb[SHEET_RESUMEN])
    totales = _parse_resumen_totales(wb[SHEET_RESUMEN])

    existente = db.query(Carga).filter(Carga.fecha_extraccion == fecha).first()
    if existente:
        raise IngestError(
            f"Ya existe una carga con fecha de extracción {fecha:%d-%m-%Y %H:%M} "
            f"(carga #{existente.id}). Elimínala primero si quieres reprocesarla."
        )

    carga = Carga(
        fecha_extraccion=fecha,
        archivo_nombre=archivo_nombre,
        subido_por=usuario,
        estado="procesando",
        totales=totales,
    )
    db.add(carga)
    db.flush()  # obtener carga.id

    n_ost = _ingest_sertec(db, wb[SHEET_SERTEC], carga.id)
    n_sf = _ingest_sf(db, wb[SHEET_SF], carga.id)
    wb.close()

    carga.n_ost = n_ost
    carga.n_sf = n_sf
    carga.estado = "listo"
    db.flush()

    # Motor de alertas (import diferido para evitar ciclo de imports).
    from .diff import generar_alertas
    carga.n_alertas = generar_alertas(db, carga)

    db.commit()
    db.refresh(carga)
    return carga


def _ingest_sertec(db: Session, ws, carga_id: int) -> int:
    h = _headers(ws)

    def g(row, name):
        i = h.get(name)
        return row[i] if i is not None and i < len(row) else None

    batch, total = [], 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or g(row, "OST_NUM") is None:
            continue
        raw = {k: (row[i] if i < len(row) else None) for k, i in h.items()}
        # normaliza datetimes del raw a ISO para que sea JSON-serializable
        raw = {k: (v.isoformat() if isinstance(v, (dt.datetime, dt.date)) else v) for k, v in raw.items()}
        batch.append({
            "carga_id": carga_id,
            "ost_num": _s(g(row, "OST_NUM")),
            "f11_num": _s(g(row, "F11_NUM")),
            "f11_estado": _s(g(row, "F11_ESTADO")),
            "ost_estado": _s(g(row, "OST_ESTADO")),
            "ost_subestado": _s(g(row, "OST_SUBESTADO")),
            "ost_estado_gestion_producto": _s(g(row, "OST_ESTADO_GESTION_PRODUCTO")),
            "sub_estado_externo": _s(g(row, "SUB_ESTADO_EXTERNO")),
            "flag_plazo": _s(g(row, "FLAG_PLAZO")),
            "dias_sertec": _as_int(g(row, "DIAS_SERTEC")),
            "rango_sertec": _s(g(row, "RANGO_SERTEC")),
            "flag_responsable": _s(g(row, "FLAG_RESPONSABLE")),
            "responsable_sac": _s(g(row, "RESPONSABLE_SAC")),
            "responsable_mini_big_ticket": _s(g(row, "RESPONSABLE_MINI_BIG_TICKET")),
            "fecha_creacion": _as_dt(g(row, "OST_FECHA_CREACION")),
            "fecha_compromiso": _as_dt(g(row, "OST_FECHA_COMPROMISO")),
            "fecha_cierre": _as_dt(g(row, "OST_FECHA_CIERRE")),
            "prod_nombre": _s(g(row, "PROD_NOMBRE")),
            "prod_sku": _s(g(row, "PROD_SKU")),
            "prod_marca": _s(g(row, "PROD_MARCA")),
            "prod_modelo": _s(g(row, "PROD_MODELO")),
            "prod_garantia_modalidad": _s(g(row, "PROD_GARANTIA_MODALIDAD")),
            "prod_tipo_garantia": _s(g(row, "PROD_TIPO_GARANTIA")),
            "desc_falla": _s(g(row, "OST_DESC_FALLA")),
            "cruce_tienda": _s(g(row, "CRUCE_TIENDA")),
            "codigo_tienda": _s(g(row, "CODIGO_TIENDA")),
            "region": _s(g(row, "F11SRX_REGION")),
            "comuna": _s(g(row, "F11SRX_COMUNA")),
            "desc_linea": _s(g(row, "F11SRX_DESC_LINEA")),
            "desc_sublinea": _s(g(row, "F11SRX_DESC_SUBLINEA")),
            "precio_vta": _as_float(g(row, "F11SRX_PRECIO_VTA")),
            "raw": raw,
        })
        total += 1
        if len(batch) >= 2000:
            db.bulk_insert_mappings(OstSnapshot, batch)
            batch = []
    if batch:
        db.bulk_insert_mappings(OstSnapshot, batch)
    return total


def _ingest_sf(db: Session, ws, carga_id: int) -> int:
    h = _headers(ws)

    def g(row, name):
        i = h.get(name)
        return row[i] if i is not None and i < len(row) else None

    batch, total = [], 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or g(row, "SS_NRO") is None:
            continue
        descripcion = _s(g(row, "DESCRIPCION"))
        f11, ost = sf_link.parse_descripcion(descripcion)
        raw = {k: (row[i] if i < len(row) else None) for k, i in h.items()}
        raw = {k: (v.isoformat() if isinstance(v, (dt.datetime, dt.date)) else v) for k, v in raw.items()}
        batch.append({
            "carga_id": carga_id,
            "ss_nro": _s(g(row, "SS_NRO")),
            "id_regulatorio": _s(g(row, "ID_REGULATORIO")),
            "estado": _s(g(row, "ESTADO")),
            "nivel_1": _s(g(row, "NIVEL_1")),
            "nivel_2": _s(g(row, "NIVEL_2")),
            "nivel_3": _s(g(row, "NIVEL_3")),
            "nivel_4": _s(g(row, "NIVEL_4")),
            "descripcion": descripcion,
            "fecha_creacion": _as_dt(g(row, "FECHA_CREACION")),
            "fecha_cierre": _as_dt(g(row, "FECHA_CIERRE")),
            "tipo_cierre": _s(g(row, "TIPO_CIERRE")),
            "nombre_propietario": _s(g(row, "NOMBRE_PROPIETARIO")),
            "tienda_origen": _s(g(row, "TIENDA_ORIGEN")),
            "validacion_matriz": _s(g(row, "VALIDACION_MATRIZ")),
            "motivo_no_cumple": _s(g(row, "MOTIVO_NO_CUMPLE_VALIDACION")),
            "f11_parseado": f11,
            "ost_parseada": ost,
            "link_status": sf_link.link_status(f11, ost),
            "raw": raw,
        })
        total += 1
        if len(batch) >= 1000:
            db.bulk_insert_mappings(SfSnapshot, batch)
            batch = []
    if batch:
        db.bulk_insert_mappings(SfSnapshot, batch)
    return total
