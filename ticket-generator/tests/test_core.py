"""Pruebas de la lógica de texto y Excel (no requieren navegador)."""

import io
import os
import sys

from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from ticketgen.text import build_description
from ticketgen.excel import TicketWorkbook


def test_description_con_sn():
    d = build_description("12345", "678", "910", "SN-ABC")
    assert d == (
        "OST 12345 F11 678 GD 910 SN SN-ABC "
        "tiene autorización de facturación, favor informar RMA para poder facturar."
    )


def test_description_sin_sn():
    d = build_description("12345", "678", "910", "")
    assert "SN" not in d.split("tiene")[0]
    assert d.startswith("OST 12345 F11 678 GD 910 tiene")


def _sample_xlsx(with_header=True):
    wb = Workbook()
    ws = wb.active
    if with_header:
        ws.append(["OST", "F11", "GD", "SN"])
    ws.append([12345, 678, 910, "ABC"])
    ws.append([22222, 333, 444, None])  # sin SN
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_lectura_y_escritura():
    tw = TicketWorkbook(_sample_xlsx(with_header=True), output_column="E")
    filas = tw.rows()
    assert len(filas) == 2
    assert filas[0].ost == "12345"      # sin ".0"
    assert filas[0].sn == "ABC"
    assert "SN ABC" in filas[0].description
    assert "SN" not in filas[1].description.split("tiene")[0]

    # Escribir ticket en columna E y releer.
    tw.set_ticket(filas[0].excel_row, "TCK-1")
    data = tw.to_bytes()
    tw2 = TicketWorkbook(data, output_column="E")
    assert tw2.rows()[0].ticket == "TCK-1"


def test_deteccion_encabezado():
    con = TicketWorkbook(_sample_xlsx(with_header=True))
    assert con.has_header is True
    sin = TicketWorkbook(_sample_xlsx(with_header=False))
    assert sin.has_header is False


if __name__ == "__main__":
    for name, fn in list(globals().items()):
        if name.startswith("test_") and callable(fn):
            fn()
            print("ok:", name)
    print("\nTodas las pruebas pasaron.")
