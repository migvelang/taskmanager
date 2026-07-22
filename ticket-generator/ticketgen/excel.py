"""Lectura y escritura del Excel.

Estructura de columnas esperada (fijada por el usuario):

    A -> N° de OST
    B -> N° de F11
    C -> N° de guía de despacho (GD)
    D -> N° de serie (SN)  [opcional]
    E -> N° de ticket generado  (columna de SALIDA, la escribimos nosotros)
"""

from __future__ import annotations

import io
from dataclasses import dataclass, asdict
from typing import Optional

from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string

from .text import build_description, DEFAULT_SUFFIX

# Encabezados por defecto para la columna de salida.
TICKET_HEADER = "N° Ticket"


def _cell_str(value) -> str:
    """Convierte el valor de una celda a texto limpio.

    openpyxl entrega números como int/float; queremos "12345" y no "12345.0".
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@dataclass
class TicketRow:
    excel_row: int          # número de fila real en el Excel (1-indexado)
    ost: str
    f11: str
    gd: str
    sn: str
    ticket: str             # ya existente en columna E (si lo hubiera)
    description: str        # texto generado para el ticket

    def to_dict(self) -> dict:
        return asdict(self)


class TicketWorkbook:
    """Envuelve un workbook de openpyxl y expone las filas de tickets."""

    @classmethod
    def blank(cls, output_column: str = "E", suffix: str = DEFAULT_SUFFIX) -> "TicketWorkbook":
        """Crea un Excel nuevo en blanco con encabezados A–D."""
        wb = Workbook()
        ws = wb.active
        ws.append(["OST", "F11", "GD", "SN"])
        buf = io.BytesIO()
        wb.save(buf)
        return cls(buf.getvalue(), has_header=True, output_column=output_column, suffix=suffix)

    def __init__(
        self,
        data: bytes,
        *,
        has_header: Optional[bool] = None,
        output_column: str = "E",
        suffix: str = DEFAULT_SUFFIX,
    ):
        self._wb = load_workbook(io.BytesIO(data))
        self._ws = self._wb.active
        self.output_column = output_column.upper()
        self.output_col_idx = column_index_from_string(self.output_column)
        self.suffix = suffix
        self.has_header = self._detect_header() if has_header is None else has_header
        self._ensure_output_header()

    # ---------- detección de encabezado ----------
    def _detect_header(self) -> bool:
        """Heurística: si A1 no parece un número de OST, es encabezado."""
        a1 = self._ws.cell(row=1, column=1).value
        if a1 is None:
            return False
        if isinstance(a1, (int, float)):
            return False
        # Texto en A1 -> probablemente encabezado.
        return True

    @property
    def first_data_row(self) -> int:
        return 2 if self.has_header else 1

    def _ensure_output_header(self):
        if self.has_header:
            cell = self._ws.cell(row=1, column=self.output_col_idx)
            if not _cell_str(cell.value):
                cell.value = TICKET_HEADER

    # ---------- lectura ----------
    def rows(self) -> list[TicketRow]:
        result: list[TicketRow] = []
        for r in range(self.first_data_row, self._ws.max_row + 1):
            ost = _cell_str(self._ws.cell(row=r, column=1).value)
            f11 = _cell_str(self._ws.cell(row=r, column=2).value)
            gd = _cell_str(self._ws.cell(row=r, column=3).value)
            sn = _cell_str(self._ws.cell(row=r, column=4).value)
            ticket = _cell_str(self._ws.cell(row=r, column=self.output_col_idx).value)
            # Saltar filas totalmente vacías.
            if not any([ost, f11, gd, sn]):
                continue
            desc = build_description(ost, f11, gd, sn, suffix=self.suffix)
            result.append(
                TicketRow(
                    excel_row=r,
                    ost=ost,
                    f11=f11,
                    gd=gd,
                    sn=sn,
                    ticket=ticket,
                    description=desc,
                )
            )
        return result

    # ---------- escritura ----------
    def set_ticket(self, excel_row: int, ticket_number: str) -> None:
        """Escribe el número de ticket en la columna de salida (E por defecto)."""
        self._ws.cell(row=excel_row, column=self.output_col_idx).value = ticket_number

    def row_values(self, excel_row: int) -> dict:
        """Devuelve OST/F11/GD/SN de una fila (para registrar en el historial)."""
        return {
            "ost": _cell_str(self._ws.cell(row=excel_row, column=1).value),
            "f11": _cell_str(self._ws.cell(row=excel_row, column=2).value),
            "gd": _cell_str(self._ws.cell(row=excel_row, column=3).value),
            "sn": _cell_str(self._ws.cell(row=excel_row, column=4).value),
        }

    def add_row(self, ost: str, f11: str, gd: str, sn: str = "") -> int:
        """Agrega un caso nuevo al final (columnas A–D) y devuelve su fila."""
        # Buscar la primera fila realmente vacía (evita huecos al final).
        r = self.first_data_row
        while any(
            _cell_str(self._ws.cell(row=r, column=c).value) for c in (1, 2, 3, 4)
        ):
            r += 1
        self._ws.cell(row=r, column=1).value = ost
        self._ws.cell(row=r, column=2).value = f11
        self._ws.cell(row=r, column=3).value = gd
        self._ws.cell(row=r, column=4).value = sn
        return r

    def to_bytes(self) -> bytes:
        buf = io.BytesIO()
        self._wb.save(buf)
        return buf.getvalue()
